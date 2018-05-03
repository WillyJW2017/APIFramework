# 操作Excel数据类的封装

import openpyxl

class OperationExcel:
    def __init__(self, filePath=None, sheetName=None):
        defaultfilePath = 'C:\\Willy\\Study\\API\\TestCase.xlsx'
        if filePath:
            self.filePath = filePath
            self.sheetName = sheetName
        else:
            self.filePath = defaultfilePath
            self.sheetName = 'TestCases'
        self.load_excel()

    # 加载excel，激活指定sheet
    def load_excel(self):
        self.workbook = openpyxl.load_workbook(self.filePath)
        self.sheet = self.workbook[self.sheetName]

    # 获取指定单元格内容
    def get_cell_data(self, rowNum, colNum):
        cell_data = self.sheet.cell(row=rowNum, column=colNum).value
        return cell_data

    # 获取行数
    def get_rows(self):
        row = self.sheet.max_row
        return row

    # 获取列数
    def get_cols(self):
        col = self.sheet.max_column
        return col

    # 写入数据到excel
    def write_data(self, row, col, value):
        self.sheet.cell(row=row, column=col, value=value)
        self.workbook.save(filename = self.filePath)

    # 根据case_id获取对应行的内容
    def get_row_value_by_caseId(self, case_id):
        row_num = self.get_row_num_by_caseId(case_id)
        rows_value = self.get_row_value_by_rowNum(row_num)
        return rows_value

    # 根据case_id获取对应的行号
    def get_row_num_by_caseId(self, case_id):
        num = 1
        cols_value = self.get_col_value_by_colId()
        for col_value in cols_value:
            if col_value.value in case_id:
                return num
            num = num + 1

    # 根据行号找到该行的内容
    def get_row_value_by_rowNum(self, row):
        rowValue = self.sheet[row]
        return rowValue

    # 根据列的id(A,B,C...)，获取某一列的内容，默认获取第一列
    def get_col_value_by_colId(self, col_id = None):
        if col_id != None:
            colData = self.sheet[col_id]
        else:
            colData = self.sheet['A']
        return colData

# if __name__ == '__main__':
# #     # op = operationExcel('C:\\Willy\\Study\\API\\test.xlsx', 'UFTConfig')
#     op = OperationExcel()
# #     # print(op.get_cell_data(4, 1))
# #     # print(op.get_lines())
# #     op.write_data(10,1,'PASS!!!')
#     print(op.get_row_value_by_caseId('TC02'))
